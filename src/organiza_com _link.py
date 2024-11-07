import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging
import os

from main import connect_to_salesforce  # Certifique-se de que essa função de conexão esteja definida corretamente

# Configuração de logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Arquivo de entrada e saída
input_file = 'data_source/conta_Somente_duplicadas.xlsx'  # Arquivo de origem na pasta data_source
output_file = 'data/contas_duplicadas_com_cores.xlsx'  # Arquivo de saída com cores
debug_folder = 'data/debug/'  # Pasta para armazenar arquivos de debug

# Cria a pasta de debug, se não existir
if not os.path.exists(debug_folder):
    os.makedirs(debug_folder)

def load_duplicate_ids(filename):
    """Carrega os IDs das contas duplicadas a partir de um arquivo Excel."""
    try:
        df_ids = pd.read_excel(filename)
        logger.debug(f"IDs carregados: {df_ids['Id'].dropna().tolist()}")  # Adicionando log para os IDs
        return df_ids['Id'].dropna().tolist()  # Agora utilizando o campo 'Id' corretamente
    except Exception as e:
        logger.error(f"Erro ao carregar IDs duplicados: {e}")
        raise

def fetch_accounts_by_ids(sf, ids_list):
    """Consulta todas as contas no Salesforce para os IDs especificados e retorna os registros."""
    try:
        fields = ', '.join([f['name'] for f in sf.Account.describe()['fields']])
        query = f"SELECT {fields} FROM Account WHERE Id IN ({', '.join([f"'{id}'" for id in ids_list])})"
        logger.info("Buscando contas duplicadas no Salesforce...")
        result = sf.query_all(query)['records']
        
        # Limitar a quantidade de contas para log (mostra as 5 primeiras)
        logger.debug(f"Contas retornadas do Salesforce (exibindo as 5 primeiras): {result[:5]}")
        
        return result
    except Exception as e:
        logger.error(f"Erro ao buscar contas: {e}")
        raise

def count_non_empty_fields(account):
    """Conta quantos campos não estão vazios em uma conta."""
    return sum(1 for value in account.values() if value not in [None, '', 'NA'])

def choose_base_account(accounts, group_key):
    """Escolhe a conta base (a que será mantida) entre várias contas duplicadas."""
    logger.debug(f"Contas a serem comparadas: {accounts}")  # Log para ver as contas passadas

    if not accounts:
        logger.warning(f"Nenhuma conta para comparar no grupo {group_key}!")
        return None, []

    base_account = accounts[0]
    discarded_account_ids = []  # IDs das contas descartadas

    # Ordenar contas pela data de modificação (mais recente primeiro)
    accounts = sorted(accounts, key=lambda x: x.get('LastModifiedDate', ''), reverse=True)

    logger.debug(f"Contas ordenadas para o grupo {group_key}: {accounts}")  # Log para verificar as contas ordenadas

    for account in accounts[1:]:
        last_modified_base = base_account.get('LastModifiedDate')
        last_modified_account = account.get('LastModifiedDate')

        logger.debug(f"Comparando conta base (ID {base_account['Id']}) e conta duplicada (ID {account['Id']})")

        # Pausar para ver as informações antes de continuar
        input(f"Comparando conta base (ID {base_account['Id']}) com conta duplicada (ID {account['Id']}). Pressione Enter para continuar...")

        # Se a data de modificação da conta atual for mais recente que a da conta base, substitui
        if last_modified_account > last_modified_base:
            discarded_account_ids.append(base_account['Id'])
            base_account = account  # Atualiza a conta base
            logger.debug(f"Conta base foi atualizada para: {base_account['Id']}")

        elif last_modified_account == last_modified_base:
            # Se as datas forem iguais, escolhe a conta com mais campos preenchidos
            count_base = count_non_empty_fields(base_account)
            count_account = count_non_empty_fields(account)
            
            logger.debug(f"Contagem de campos preenchidos na conta base: {count_base}, conta duplicada: {count_account}")

            # Pausar para ver as informações antes de continuar
            input(f"Contagem de campos preenchidos na conta base ({count_base}) e conta duplicada ({count_account}). Pressione Enter para continuar...")

            if count_account > count_base:
                discarded_account_ids.append(base_account['Id'])
                base_account = account  # Atualiza a conta base
                logger.debug(f"Conta base foi atualizada para: {base_account['Id']}")

        # Agora, preenche os campos vazios da base com os dados da conta secundária
        for field in base_account.keys():
            value_base = base_account.get(field)
            value_account = account.get(field)
            
            # Se o campo na base estiver vazio ou for 'NA', usa o valor da conta secundária
            if value_base in [None, '', 'NA']:
                base_account[field] = value_account  # Preenche com o valor da conta duplicada

    base_account['Discarded_Account_Ids'] = discarded_account_ids  # Adiciona os IDs das contas descartadas
    
    # Salva o grupo de contas em um arquivo de texto para depuração
    debug_filename = os.path.join(debug_folder, f"group_{group_key}_debug.txt")
    with open(debug_filename, 'w') as f:
        f.write(f"Grupo: {group_key}\n")
        f.write(f"Contas comparadas:\n")
        for account in accounts:
            f.write(f"Conta ID {account['Id']} - Última Modificação: {account.get('LastModifiedDate')}\n")
        f.write(f"Conta base escolhida: {base_account['Id']}\n")
        f.write(f"Contas descartadas: {discarded_account_ids}\n")

    return base_account, discarded_account_ids  # Retorna a conta base e os IDs descartados

def consolidate_accounts(df):
    """Consolida várias contas duplicadas em uma única conta base e aplica a formatação de cores."""
    consolidated_accounts = []

    # Limpeza dos valores nas colunas de agrupamento
    logger.debug("Limpando valores nas colunas NEO_Cpfcnpj__c e NEO_Clave_Cliente__c antes do agrupamento...")
    df['NEO_Cpfcnpj__c'] = df['NEO_Cpfcnpj__c'].str.strip()  # Remover espaços extras
    df['NEO_Clave_Cliente__c'] = df['NEO_Clave_Cliente__c'].str.strip()  # Remover espaços extras

    # Verifique os valores únicos nas colunas de agrupamento
    logger.debug(f"Valores únicos em NEO_Cpfcnpj__c após limpeza: {df['NEO_Cpfcnpj__c'].unique()}")
    logger.debug(f"Valores únicos em NEO_Clave_Cliente__c após limpeza: {df['NEO_Clave_Cliente__c'].unique()}")

    # Agrupar contas por NEO_Cpfcnpj__c e NEO_Clave_Cliente__c
    grouped_accounts = df.groupby(['NEO_Cpfcnpj__c', 'NEO_Clave_Cliente__c'])

    logger.debug(f"Grupos de contas a serem consolidadas: {grouped_accounts}")  # Log para verificar o agrupamento

    for group_key, group in grouped_accounts:
        accounts = group.to_dict('records')
        logger.debug(f"Grupo de contas: {accounts}")  # Log para verificar os grupos
        base_account, discarded_ids = choose_base_account(accounts, group_key)

        if base_account is None:
            logger.warning(f"Sem conta base para o grupo: {group_key}")
            continue

        # Adiciona a conta base à lista de contas consolidadas
        consolidated_accounts.append(base_account)

    # Converter para DataFrame final
    df_consolidated = pd.DataFrame(consolidated_accounts)

    return df_consolidated

def save_to_excel(df):
    """Aplica a formatação de cores no arquivo Excel."""
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Contas Consolidadas')

        wb = load_workbook(output_file)
        ws = wb['Contas Consolidadas']

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Aplica a formatação de cores
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            conta_id = row[0].value  # ID da conta
            discarded = row[0].offset(0, 16).value  # Coluna de contas descartadas (baseada no índice)
            if discarded:  # Se a conta foi descartada
                row[0].fill = yellow_fill
            else:  # Se for a conta base
                row[0].fill = green_fill

        wb.save(output_file)
        logger.info(f"Arquivo com contas consolidadas salvo em {output_file}")

    except Exception as e:
        logger.error(f"Erro ao salvar arquivo Excel: {e}")
        raise

if __name__ == "__main__":
    try:
        # Carregar os IDs das contas duplicadas
        ids_list = load_duplicate_ids(input_file)
        
        # Conectar ao Salesforce
        sf = connect_to_salesforce()

        # Buscar contas duplicadas
        accounts = fetch_accounts_by_ids(sf, ids_list)
        
        # Consolidar contas duplicadas
        consolidated_df = consolidate_accounts(pd.DataFrame(accounts))
        
        # Salvar o resultado com cores
        save_to_excel(consolidated_df)
    except Exception as e:
        logger.error(f"Erro no processo: {e}")
