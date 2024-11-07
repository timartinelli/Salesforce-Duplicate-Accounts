import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from main import connect_to_salesforce  # Função de conexão importada do arquivo main.py

# Configuração de logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Arquivo de entrada e saída
input_file = 'data_source/conta_Somente_duplicadas.xlsx'  # Arquivo de origem na pasta data_source
output_file = 'data/contas_duplicadas_consolidadas_com_cores.xlsx'  # Arquivo de saída com cores

def load_duplicate_ids(filename):
    """Carrega os IDs das contas duplicadas a partir de um arquivo Excel."""
    try:
        df_ids = pd.read_excel(filename)
        return df_ids['Id'].dropna().tolist()  # Agora utilizando o campo 'Id' corretamente
    except Exception as e:
        logger.error(f"Erro ao carregar IDs duplicados: {e}")
        raise

def fetch_accounts_by_ids(sf, ids_list):
    """Consulta todas as contas no Salesforce para os IDs especificados e retorna os registros."""
    try:
        fields = ', '.join([f['name'] for f in sf.Account.describe()['fields']])
        query = f"SELECT {fields} FROM Account WHERE Id IN ({', '.join([f"'{id}'" for id in ids_list])})"
        logger.info("Buscando contas duplicadas no Salesforce...")
        return sf.query_all(query)['records']
    except Exception as e:
        logger.error(f"Erro ao buscar contas: {e}")
        raise

def count_non_empty_fields(account):
    """Conta quantos campos não estão vazios em uma conta."""
    return sum(1 for value in account.values() if value not in [None, '', 'NA'])

def choose_base_account(accounts):
    """Escolhe a conta base (a que será mantida) entre várias contas duplicadas."""
    base_account = accounts[0]
    discarded_account_ids = []  # IDs das contas descartadas

    # Ordenar contas pela data de modificação (mais recente primeiro)
    accounts = sorted(accounts, key=lambda x: x.get('LastModifiedDate', ''), reverse=True)

    for account in accounts[1:]:
        last_modified_base = base_account.get('LastModifiedDate')
        last_modified_account = account.get('LastModifiedDate')

        # Se a data de modificação da conta atual for mais recente que a da conta base, substitui
        if last_modified_account > last_modified_base:
            discarded_account_ids.append(base_account['Id'])
            base_account = account  # Atualiza a conta base
        elif last_modified_account == last_modified_base:
            # Se as datas forem iguais, escolhe a conta com mais campos preenchidos
            count_base = count_non_empty_fields(base_account)
            count_account = count_non_empty_fields(account)
            
            if count_account > count_base:
                discarded_account_ids.append(base_account['Id'])
                base_account = account  # Atualiza a conta base

        # Agora, preenche os campos vazios da base com os dados da conta secundária
        for field in base_account.keys():
            value_base = base_account.get(field)
            value_account = account.get(field)
            
            # Se o campo na base estiver vazio ou for 'NA', usa o valor da conta secundária
            if value_base in [None, '', 'NA']:
                base_account[field] = value_account  # Preenche com o valor da conta duplicada

    base_account['Discarded_Account_Ids'] = discarded_account_ids  # Adiciona os IDs das contas descartadas
    return base_account, discarded_account_ids  # Retorna a conta base e os IDs descartados

def consolidate_accounts(df):
    """Consolida várias contas duplicadas em uma única conta base e aplica a formatação de cores."""
    consolidated_accounts = []
    discarded_account_ids = []

    # Agrupar por ID da conta
    grouped_accounts = df.groupby('Id')

    for account_id, group in grouped_accounts:
        accounts = group.to_dict('records')
        base_account, discarded_ids = choose_base_account(accounts)

        # Adiciona a conta base à lista de contas consolidadas
        consolidated_accounts.append(base_account)

        # Adiciona os IDs descartados à lista geral
        discarded_account_ids.extend(discarded_ids)

    # Converter para DataFrame final
    df_consolidated = pd.DataFrame(consolidated_accounts)

    return df_consolidated

def apply_formatting_to_excel(df):
    """Aplica a formatação de cores no arquivo Excel, marcando contas consolidadas em verde e duplicadas em amarelo."""
    # Carregar o arquivo Excel para edição
    wb = load_workbook(output_file)
    ws = wb.active

    # Definir os estilos de cores
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Loop para percorrer as contas e aplicar as cores
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        account_id_cell = row[0]  # Assumindo que o ID da conta esteja na primeira coluna
        discarded_id_cell = row[6]  # A coluna 'Discarded_Account_Ids' é a 7ª (index 6)

        if discarded_id_cell.value == "":  # Conta não foi descartada
            account_id_cell.fill = green_fill  # Conta base em verde
        else:
            account_id_cell.fill = yellow_fill  # Conta duplicada em amarelo

    # Salvar o arquivo com a formatação aplicada
    wb.save(output_file)
    logger.info(f"Contas consolidadas com cores salvas em: {output_file}")

def main():
    """Função principal que conecta ao Salesforce, carrega duplicatas e realiza a consolidação automática."""
    try:
        # Conectar ao Salesforce
        sf = connect_to_salesforce()
        
        # Carregar IDs das contas duplicadas
        ids_list = load_duplicate_ids(input_file)
        logger.info(f"{len(ids_list)} IDs duplicados carregados para comparação.")
        
        # Buscar contas duplicadas no Salesforce
        accounts = fetch_accounts_by_ids(sf, ids_list)
        df_accounts = pd.DataFrame(accounts)

        # Consolidação das contas duplicadas
        df_consolidated = consolidate_accounts(df_accounts)
        
        # Salvar o resultado consolidado em um arquivo Excel
        df_consolidated.to_excel(output_file, index=False)
        logger.info(f'Contas duplicadas consolidadas salvas em: {output_file}')
        
        # Aplicar a formatação de cores no arquivo Excel
        apply_formatting_to_excel(df_consolidated)
        
    except Exception as e:
        logger.error(f"Ocorreu um erro na execução: {e}")

if __name__ == "__main__":
    main()
